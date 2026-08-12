#!/usr/bin/env python3
"""Drain queue.jsonl into my-vocab-sheet.xlsx. Pure code — no model, no network.

    ./flush_queue.py [--dry] [--rebuild] [--sheet PATH] [--queue PATH]

The Stop hook writes one JSON object per item into the queue; this turns each into a
sheet row. Rows go in at row 3, directly under the config row, so the newest are on top.

Consumed items are appended to `history.jsonl` and never deleted. That file is the
record; the workbook is only a rendering of it, and it is the fragile half — any
spreadsheet app holding it open will save its own in-memory copy over ours and take rows
with it. Deleting the history after a sync would make the next such accident permanent,
which is exactly the accident it exists for.

**The workbook is rewritten in full on every run, never patched.** Patching would assume
the file on disk is the one we last wrote, and it is not: what comes back may be an older
copy whose IDs predate the current scheme, in which case nothing matches and every row is
added a second time. Re-rendering assumes nothing, so a clobbered, stale or missing
workbook all repair themselves without a special case.

Each item carries `synced`, the moment it first reached the sheet. `--rebuild` differs
only in ignoring the queue.

Every cell comes straight from what the answer printed, so rows arrive finished and no
model is involved anywhere in the path.
"""
import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook

VOCAB_DIR = os.path.dirname(os.path.abspath(__file__))
SHEET = os.path.join(VOCAB_DIR, "my-vocab-sheet.xlsx")
QUEUE = os.path.join(VOCAB_DIR, "queue.jsonl")
HISTORY = os.path.join(VOCAB_DIR, "history.jsonl")    # append-only, the real record

# The short part-of-speech the answer uses -> the word the ID and TAGS spell out.
POS_LONG = {
    "n": "noun", "v": "verb", "adj": "adjective", "adv": "adverb",
    "prep": "preposition", "conj": "conjunction", "pron": "pronoun",
    "phr v": "phrasal-verb", "idiom": "idiom",
}


def slug(text):
    return re.sub(r"-+", "-", re.sub(r"[^\w]+", "-", text.lower())).strip("-")


def clean(text):
    """Sheet cells are fetched as TSV, so a tab would split the row in half."""
    return text.replace("\t", " ").replace("\r", " ").strip()


def html(text):
    return clean(text).replace('"', "&quot;")


def strip_md(text):
    return re.sub(r"\*{1,2}([^*]+)\*{1,2}", r"\1", text)


def vocab_row(item):
    # One row per word, forever: meeting the word again refreshes the card he already
    # has rather than opening a second one.
    pos = POS_LONG.get(item.get("pos", "").lower(), item.get("pos", "")).strip()
    return {
        "ID": f"{slug(item['word'])}-{slug(pos)}",
        "SYNC": "TRUE",
        "SUBDECK 1": item["date"],
        "Word": clean(item["word"]),
        "IPA": clean(item.get("ipa", "")),
        "POS": pos,
        "Meaning": clean(item.get("meaning", "")),
        "Collocation": clean(item.get("collocation", "")),
        "Example": html(item.get("example", "")),
    }


def grammar_row(item):
    digest = hashlib.sha1(item["original"].encode("utf-8")).hexdigest()[:8]
    fixes = item.get("fixes", [])
    return {
        "ID": f"g-{digest}",
        "SYNC": "TRUE",
        "SUBDECK 1": item["date"],
        "Original": html(item["original"]),
        "Corrected": html(item["corrected"]),
        "Fixes": "<br>".join(
            html(strip_md(f"[{f['tag']}] {f['body']}")) for f in fixes),
    }


def build_rows(items):
    """History -> the complete row set for both tabs, newest first.

    A word met again is not a second row and not an edit to an old one either: the
    latest capture replaces it outright and the row moves to the day it was met again.
    Repeats are the words worth revisiting, so they belong at the top of today rather
    than buried in whichever day happened to be the first.
    """
    built = {"vocab": [], "grammar": []}
    refreshed = skipped = 0
    for tab, builder in (("vocab", vocab_row), ("grammar", grammar_row)):
        index = {}
        for item in items:
            if item.get("kind", "vocab") != tab:
                continue
            try:
                row = builder(item)
            except KeyError:
                continue
            seen = index.get(row["ID"])
            if seen is not None:
                if tab != "vocab":
                    skipped += 1     # the same sentence corrected twice
                    continue
                # Drop it here and append below: history is walked oldest first, so the
                # row comes back at the end, which `reverse()` turns into the top.
                built[tab].remove(seen)
                refreshed += 1
            index[row["ID"]] = row
            built[tab].append(row)
        built[tab].reverse()         # newest on top
    return built, refreshed, skipped


def read_items(path):
    items = []
    try:
        with open(path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    items.append(json.loads(line))
                except ValueError:
                    pass
    except OSError:
        pass
    return items


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry", action="store_true", help="report, write nothing")
    ap.add_argument("--rebuild", action="store_true",
                    help="replay history.jsonl — repairs a sheet another app overwrote")
    ap.add_argument("--sheet", default=SHEET)
    ap.add_argument("--queue", default=QUEUE)
    args = ap.parse_args()

    items = read_items(HISTORY)
    pending = [] if args.rebuild else read_items(args.queue)
    replayed = len(items)
    items += pending
    if not items:
        print("nothing to flush")
        return 0

    # Both tabs are rewritten from history every run, never patched row by row.
    # Patching assumed the workbook on disk was the one we last wrote — and it is not:
    # an editor holding it open puts back an older copy, whose IDs may predate the
    # current scheme, so nothing matches and every row gets added a second time. A full
    # render has no such assumption. The sheet is a view of history; make it that view.
    built, refreshed, skipped = build_rows(items)

    if args.dry:
        for tab, rows in built.items():
            for row in rows:
                print(f"{tab}: {row['ID']}")
        print(f"would add {sum(len(r) for r in built.values())}, skip {skipped}")
        return 0

    # A fresh workbook every time, so a clobbered or half-deleted one needs no repair
    # path of its own: there is nothing to reconcile with, only a view to re-render.
    sys.path.insert(0, VOCAB_DIR)
    import gen_sheet as g

    wb = Workbook()
    ws = wb.active
    ws.title = "vocab"
    g.fill(ws, g.HEADERS, g.CONFIG_ROW, g.COLUMN_WIDTHS, built["vocab"])
    g.fill(wb.create_sheet("grammar"), g.G_HEADERS, g.G_CONFIG_ROW,
           g.G_COLUMN_WIDTHS, built["grammar"])
    wb.save(args.sheet)
    added = sum(len(rows) for rows in built.values())

    # Report what is on disk, not what was attempted. openpyxl returning cleanly only
    # means the bytes were written; another program holding the workbook open can save
    # its own older copy straight over them, and that leaves no error to catch. Reading
    # the file back is the only statement about it that cannot be wrong.
    on_disk = 0
    try:
        check = load_workbook(args.sheet)
        on_disk = sum(check[t].max_row - 2 for t in check.sheetnames
                      if t in ("vocab", "grammar"))
    except Exception as exc:
        print(f"could not verify the sheet: {exc}", file=sys.stderr)
    expected = sum(len(rows) for rows in built.values())
    if on_disk < expected:
        print(f"WARNING: verified {on_disk} row(s) on disk, expected {expected} — "
              f"another program is overwriting {os.path.basename(args.sheet)}",
              file=sys.stderr)

    stamp = datetime.now().isoformat(timespec="seconds")
    for item in items:
        item.setdefault("synced", stamp)   # keep the first sync time, not the latest

    if replayed:
        # History was read back in, so rewrite those lines in place rather than
        # appending them a second time. Only `pending` is new.
        with open(HISTORY, "w", encoding="utf-8") as fh:
            for item in items[:replayed]:
                fh.write(json.dumps(item, ensure_ascii=False) + "\n")

    if pending:
        # Append to history first, drop the queue second: a crash between the two costs
        # a duplicate line, which the ID dedup eats. The other order would cost a word.
        with open(HISTORY, "a", encoding="utf-8") as fh:
            for item in items[replayed:]:
                fh.write(json.dumps(item, ensure_ascii=False) + "\n")
    if not args.rebuild and os.path.exists(args.queue):
        os.remove(args.queue)

    print(f"added {added}, refreshed {refreshed}, skipped {skipped}, "
          f"verified {on_disk} on disk")
    return 0


if __name__ == "__main__":
    sys.exit(main())
