#!/usr/bin/env python3
"""Create an empty my-vocab sheet in the exact shape sheets2anki expects.

    ./gen_sheet.py my-vocab-sheet.xlsx [--sample]

Two tabs, same contract, different subject: `vocab` holds words, `grammar` holds the
corrections made to Tân's own sentences. sheets2anki is pointed at one tab at a time
(the `gid` in the URL), so the second tab is a second deck, never a contamination of
the first.

Writes row 1 (headers) and row 2 (the #config directive row) and stops. Data rows are
added by an agent, newest first, starting at row 3 — see SHEET-FORMAT.md.

Running this on an existing file OVERWRITES it, so it is a bootstrap tool, not an
appender. Regenerating a sheet that already holds words would throw those words away.
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Left-to-right order IS the card order, so this list is the card design.
# Only ID / SYNC / SUBDECK n / TAGS are reserved; every other header becomes an Anki field.
HEADERS = [
    "ID",           # permanent key, <word>-<pos>: one row per word, forever
    "SYNC",         # per-row on/off switch
    "SUBDECK 1",    # the day, e.g. 2026-08-11 -> one subdeck per day
    "Word",         # first content column, so it is on the front by default
    "IPA",
    "POS",
    "Meaning",
    "Collocation",
    "Example",
]

# Row 2. Each cell is "key=value; key=value" for that column; an empty cell means defaults.
CONFIG_ROW = {
    "ID": "#config align=center",
    "Word": "side=front; size=40; bold; tts=en_US; voices=Apple_Ava_(Premium)",
    "IPA": "side=front; size=18; color=muted",
    "POS": "side=front; size=14; color=accent; italic",
    "Meaning": "size=22",
    "Collocation": "label=Collocations; size=17; color=accent",
    "Example": "label=Examples; size=17; italic; tts=en_US; voices=Apple_Ava_(Premium)",
}

SAMPLE = [
    {
        "ID": "postpone-verb",
        "SYNC": "TRUE",
        "SUBDECK 1": "2026-08-11",
        "Word": "postpone",
        "IPA": "/poʊstˈpoʊn/",
        "POS": "verb",
        "Meaning": "To delay or put off an event, appointment etc.",
        "Collocation": "postpone a release · postpone the golive · postpone indefinitely",
        "Example": "We had to postpone the golive until the pipeline went green.<br>"
                   "They postponed the review because the spec was still changing.",
    },
]

COLUMN_WIDTHS = {"A": 22, "B": 7, "C": 13, "D": 16, "E": 16, "F": 8,
                 "G": 44, "H": 44, "I": 52}

# ---------------------------------------------------------------- grammar tab

# The front is what Tân wrote, the back is the fixed version — so the card asks him to
# produce the correction he was given, which is the only part of feedback that sticks.
G_HEADERS = [
    "ID",           # g-<sha1[:8] of Original>
    "SYNC",
    "SUBDECK 1",    # the day he was corrected
    "Original",     # first content column -> front
    "Corrected",
    "Fixes",        # the tagged bullets, joined by <br>
]

G_CONFIG_ROW = {
    "ID": "#config align=center",
    "Original": "side=front; size=26; italic; color=muted",
    "Corrected": "size=26; bold; tts=en_US",
    "Fixes": "label=Fixes; size=16; color=accent",
}

G_SAMPLE = [
    {
        "ID": "g-1f4a9c02",
        "SYNC": "TRUE",
        "SUBDECK 1": "2026-08-11",
        "Original": "dose this work for practice my english?",
        "Corrected": "Does this work for practicing my English?",
        "Fixes": "[Spelling] &quot;dose&quot; → &quot;does&quot;<br>"
                 "[Grammar] &quot;for practice&quot; → &quot;for practicing&quot; — "
                 "after &quot;for&quot;, use the -ing form",
    },
]

G_COLUMN_WIDTHS = {"A": 22, "B": 7, "C": 13, "D": 58, "E": 58, "F": 70}


def fill(ws, headers, config_row, widths, sample):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="DDE7F0")

    ws.append([config_row.get(h, "") for h in headers])
    for cell in ws[2]:
        cell.font = Font(italic=True, color="7A7A7A")

    for row in sample:
        ws.append([row.get(h, "") for h in headers])

    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    out = args[0] if args else "my-vocab-sheet.xlsx"
    want_sample = "--sample" in sys.argv[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "vocab"
    fill(ws, HEADERS, CONFIG_ROW, COLUMN_WIDTHS, SAMPLE if want_sample else [])
    fill(wb.create_sheet("grammar"), G_HEADERS, G_CONFIG_ROW, G_COLUMN_WIDTHS,
         G_SAMPLE if want_sample else [])

    wb.save(out)
    print(f"created: {out} (tabs: vocab, grammar)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
